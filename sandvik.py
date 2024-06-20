from bs4 import BeautifulSoup
from selenium import webdriver
import re
import time
import requests
from openpyxl import load_workbook
import os


def get_tools():
    nos = []
    i = 4
    while True:
        if ws[f"d{i}"].value:
            nos.append(str(ws[f"d{i}"].value).replace(" ", ""))
        if i > 10000:
            print(len(nos))
            break
        i += 1
    return nos


def scrape_and_download(art_no, link):
    site = link
    driver = webdriver.Chrome()
    driver.get(site)
    time.sleep(2)

    pics = BeautifulSoup(driver.page_source, 'html.parser')
    img_tags = pics.find_all('img')

    urls = [img['src'] for img in img_tags]
    found = False

    for url in urls:
        filename = \
            re.search(r'^https://productinformation.sandvik.coromant.com/s3/documents/pictures/pict-3d-view', url)
        if filename:
            with open(f'sandvik/{art_no}.gif', 'wb') as handle:
                response = requests.get(url, stream=True)
                if not response.ok:
                    print(response)
                for block in response.iter_content(1024):
                    if not block:
                        break

                    handle.write(block)
            found = True
    if not found:
        print(f"Did not find picture for {art_no}, broken link? {link}")


def gen_link(art_no):
    return f"https://www.sandvik.coromant.com/en-gb/product-details?m={art_no}"


wb = load_workbook('sandvik.xlsx', data_only=True)
ws = wb["Sheet1"]

art_nos = get_tools()

for art in art_nos:
    # if os.path.exists(f"sandvik/{art}.gif"):
    #     print(f"Already downloaded picture for {art}")
    if not os.path.exists(f"sandvik/{art}.gif"):
        scrape_and_download(art, gen_link(art))
