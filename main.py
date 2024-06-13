from bs4 import BeautifulSoup
from selenium import webdriver
import re
import time
import requests
from openpyxl import load_workbook


def get_tools():
    nos = {}
    i = 2
    while True:
        if ws[f"a{i}"].value:
            nos.update({ws[f"a{i}"].value: ws[f"r{i}"].value})
        if not ws[f"a{i}"].value:
            break
        i += 1
    return nos


def scrape_and_download(art, link):
    site = link
    driver = webdriver.Chrome()
    driver.get(site)
    time.sleep(3)

    pics = BeautifulSoup(driver.page_source, 'html.parser')
    img_tags = pics.find_all('img')

    urls = [img['src'] for img in img_tags]

    for url in urls:
        filename = \
            re.search(r'^https://productinformation.sandvik.coromant.com/s3/documents/pictures/pict-3d-view', url)
        if filename:
            print(f"Found picture for: {art}, downloading...")
            with open(f'sandvik/{art}.gif', 'wb') as handle:
                response = requests.get(url, stream=True)
                if not response.ok:
                    print(response)
                for block in response.iter_content(1024):
                    if not block:
                        break

                    handle.write(block)
        if not filename:
            print(f"Could not scrape picture for {art}")


wb = load_workbook('Coroturn_CXS.xlsx', data_only=True)
ws = wb["Ark1"]

art_nos = get_tools()

for key in art_nos:
    scrape_and_download(key, art_nos[key])
