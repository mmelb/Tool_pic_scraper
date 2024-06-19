from bs4 import BeautifulSoup
from selenium import webdriver
import re
import time
import requests
from openpyxl import load_workbook
import os


def get_tools():
    items = []
    i = 4
    while True:
        if ws[f"d{i}"].value and not os.path.exists(f"guhring/{ws[f'd{i}'].value}.gif"):
            items.append(ws[f'd{i}'].value)
        if not ws[f"d{i}"].value:
            break
        i += 1
    return items


def scrape_and_download(art, link, line):
    site = link
    driver = webdriver.Chrome()
    driver.get(site)
    time.sleep(2)

    pics = BeautifulSoup(driver.page_source, 'html.parser')
    img_tags = pics.find_all('img')

    urls = [img['src'] for img in img_tags]

    for url in urls:
        filename = \
            re.search(rf'{line}.jpg$', url)
        if filename:
            # print(f"Found picture for: {art}, downloading...")
            with open(f'guhring/{art}.gif', 'wb') as handle:
                response = requests.get(url, stream=True)
                if not response.ok:
                    print(response)
                for block in response.iter_content(1024):
                    if not block:
                        break

                    handle.write(block)
            break


def gen_link(item):
    line = item.split(" ")[0]
    dim = item.split(" ")[1]
    formatted_dim = dim.split(",")[0].zfill(3) + dim.split(",")[1] + ((4 - len(dim.split(",")[1])) * "0")
    formatted_line = str(line).zfill(4)
    return f"https://webshop.guehring.no/0000090{formatted_line}{formatted_dim}"


wb = load_workbook('guhring.xlsx', data_only=True)
ws = wb["Sheet1"]

tools = get_tools()
print(f"Attempting to scrape pictures for {len(tools)} tools.")

failed = 0
for tool in tools:
    scrape_and_download(tool, gen_link(tool), tool.split(" ")[0])
    if not os.path.exists(f"guhring/{tool}.gif"):
        print(f"Download failed for {tool}: {gen_link(tool)}")
        failed += 1

print(f"Failed to scrape {failed} tools.")
