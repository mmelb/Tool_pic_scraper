from bs4 import BeautifulSoup
from selenium import webdriver
import re
import time
import requests
from openpyxl import load_workbook
import os


def get_tools():
    items = []
    i = 5
    while True:
        if ws[f"d{i}"].value and not os.path.exists(f"kennametal/{ws[f'd{i}'].value}.gif"):
            items.append(ws[f'd{i}'].value)
        if not ws[f"d{i}"].value:
            break
        i += 1
    return items


def gen_link(item):
    return f"https://www.kennametal.com/us/en/products/p.{item}"


def scrape_and_download(art, link):
    site = link
    driver = webdriver.Chrome()
    driver.get(site)
    time.sleep(2)

    pics = BeautifulSoup(driver.page_source, 'html.parser')
    img_tags = pics.find_all('img')

    urls = [img['src'] for img in img_tags]
    count = 0
    for url in urls:
        filename = \
            re.search(
                r'^https://images.kennametal.com/is/image/Kennametal/', url)
        if filename:
            count += 1
            if count == 1:  # Change this if wrong picture is downloaded.
                with open(f'kennametal/{art}.gif', 'wb') as handle:
                    response = requests.get(url, stream=True)
                    if not response.ok:
                        print(response)
                    for block in response.iter_content(1024):
                        if not block:
                            break

                        handle.write(block)
                break


wb = load_workbook('kennametal.xlsx', data_only=True)
ws = wb["Sheet1"]

tools = get_tools()
print(f"Attempting to scrape pictures for {len(tools)} tools.")

failed = 0
for tool in tools:
    # print(tool)
    # print(gen_link(tool))
    scrape_and_download(tool, gen_link(tool))
    if not os.path.exists(f"kennametal/{tool}.gif"):
        print(f"Download failed for {tool}: {gen_link(tool)}")
        failed += 1

print(f"Failed to scrape {failed} tools.")
