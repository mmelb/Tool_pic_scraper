from bs4 import BeautifulSoup
from selenium import webdriver
import re
import time
import requests
from openpyxl import load_workbook
import os


def get_tools():
    items = []
    i = 4
    while True:
        if ws[f"d{i}"].value and not os.path.exists(f"seco/{ws[f'd{i}'].value}.gif"):
            items.append(ws[f'd{i}'].value)
        if not ws[f"d{i}"].value:
            break
        i += 1
    return items


def gen_link(item):
    if item[3] == "-":
        item = item.split("-")[1]
    if item[3] == " ":
        item = item.split(" ")[1]
    return f"https://www.secotools.com/article/p_{item.zfill(8)}?entryPoint=ProductDetails%2FWS"


def scrape_and_download(art, link):
    site = link
    driver = webdriver.Chrome()
    driver.get(site)
    time.sleep(2)

    pics = BeautifulSoup(driver.page_source, 'html.parser')
    img_tags = pics.find_all('img')

    urls = [img['src'] for img in img_tags]

    for url in urls:
        filename = \
            re.search(
                r'^https://common-secoresources.azureedge.net/pictures/core/Content/ProductImages/As_Delivered_Image/', url)
        if filename:
            with open(f'seco/{art}.gif', 'wb') as handle:
                response = requests.get(url, stream=True)
                if not response.ok:
                    print(response)
                for block in response.iter_content(1024):
                    if not block:
                        break

                    handle.write(block)
            break


wb = load_workbook('seco.xlsx', data_only=True)
ws = wb["Sheet1"]

tools = get_tools()
print(f"Attempting to scrape pictures for {len(tools)} tools.")

failed = 0
for tool in tools:
    # print(tool)
    # print(gen_link(tool))
    scrape_and_download(tool, gen_link(tool))
    if not os.path.exists(f"seco/{tool}.gif"):
        print(f"Download failed for {tool}: {gen_link(tool)}")
        failed += 1

print(f"Failed to scrape {failed} tools.")
